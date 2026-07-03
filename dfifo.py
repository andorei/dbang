#!/usr/bin/env python3

import os
import sys
import locale
import re
import csv
import json
import glob
import argparse
import logging
import decimal as dec
from datetime import date, datetime, time
import concurrent.futures

from openpyxl import Workbook, styles, load_workbook
from openpyxl.cell import WriteOnlyCell
from jinja2 import (
    Template,
    Environment,
    FileSystemLoader,
    select_autoescape
)


VERSION = '0.4.0'

parser = argparse.ArgumentParser(
    description="Save data from input file(s) to output file(s), as specified in cfg-file specs.",
    epilog="Thanks for using %(prog)s!"
)

parser.add_argument("-v", "--version", action="version", version="%(prog)s " + VERSION)
parser.add_argument("-d", "--delete", action="store_true", help="delete processed input file(s)")
parser.add_argument("-f", "--force", action="store_true", help="process input file(s) unconditionally")
parser.add_argument("-i", "--fi", action="store", help="input file(s)")
parser.add_argument("-o", "--fo", action="store", help="output file(s)")
parser.add_argument("-u", "--user", action="store", 
                    default=os.environ.get('USER', os.environ.get('USERNAME', 'DBANG')), 
                    help="set username")
parser.add_argument("cfg_file", help="cfg-file name")
parser.add_argument("spec", nargs="?", default="all", help="spec name, defaults to \"all\"")

args = parser.parse_args()

locale.setlocale(locale.LC_TIME, '')
BASEDIR, BASENAME = os.path.split(sys.argv[0])

if not os.path.isfile(args.cfg_file) and not os.path.isfile(args.cfg_file + '.py'):
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: cfg-file not found: {args.cfg_file}\n"
    )
    sys.exit(1)

USER_DIR = os.path.expanduser('~')
if not os.path.isdir(USER_DIR):
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: user's home dir not found: {USER_DIR}\n"
    )
    sys.exit(1)
TEMP_DIR = os.path.join(USER_DIR, '.dbang')
if not os.path.isdir(TEMP_DIR):
    os.mkdir(TEMP_DIR)

CUR_DIR = os.getcwd()
CFG_DIR = os.path.abspath(os.path.dirname(args.cfg_file) or CUR_DIR)
CFG_MODULE = os.path.basename(args.cfg_file).rsplit('.', 1)[0]
sys.path.append(CFG_DIR)
cfg = __import__(CFG_MODULE)
##sources = cfg.sources
specs = cfg.specs

TEMPLATES_DIR = CFG_DIR

SPEC = args.spec
# filter out specs commented out with leading --
specs = {k:v for k,v in specs.items() if not k.startswith('--')}
if SPEC not in [*specs.keys(), 'all', *(tag for val in specs.values() if val.get('tags') for tag in val['tags'])]:
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: spec not found in cfg-file: {SPEC}\n"
    )
    sys.exit(1)

IN_DIR = getattr(cfg, 'IN_DIR', CUR_DIR)
if not os.path.isdir(IN_DIR):
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: in dir not found: {IN_DIR}\n"
    )
    sys.exit(1)

OUT_DIR = getattr(cfg, 'OUT_DIR', CUR_DIR)
if not os.path.isdir(OUT_DIR):
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: out dir not found: {OUT_DIR}\n"
    )
    sys.exit(1)

PARALLEL_WORKERS = min(getattr(cfg, 'PARALLEL_WORKERS', 1), 64)

DEBUGGING = getattr(cfg, 'DEBUGGING', False)
LOGGING = getattr(cfg, 'LOGGING', DEBUGGING)
LOG_DIR = getattr(cfg, 'LOG_DIR', CUR_DIR)
if LOGGING and not os.path.isdir(LOG_DIR):
    sys.stderr.write(
        parser.format_usage() + \
        f"{BASENAME}: error: log dir not found: {LOG_DIR}\n"
    )
    sys.exit(1)
LOG_FILE = os.path.join(LOG_DIR, f"{date.today().isoformat()}_{BASENAME.rsplit('.', 1)[0]}.log")
if sys.version_info >= (3, 9):
    logging.basicConfig(
        filename=LOG_FILE,
        encoding='utf-8',
        format="%(asctime)s:%(levelname)s:%(process)s:" + ("%(thread)d:%(message)s" if PARALLEL_WORKERS > 1 else "%(message)s"),
        level=logging.DEBUG if DEBUGGING else logging.INFO if LOGGING else logging.CRITICAL + 1
    )
else:
    logging.basicConfig(
        filename=LOG_FILE,
        format="%(asctime)s:%(levelname)s:%(process)s:" + ("%(thread)d:%(message)s" if PARALLEL_WORKERS > 1 else "%(message)s"),
        level=logging.DEBUG if DEBUGGING else logging.INFO if LOGGING else logging.CRITICAL + 1
    )
logger = logging.getLogger(__name__)

# output file encoding by default
ENCODING = getattr(cfg, 'ENCODING', locale.getpreferredencoding())
# datetime format for strftime is ISO 86101 by default
DATETIME_FORMAT = getattr(cfg, 'DATETIME_FORMAT', '%Y-%m-%d %H:%M:%S%z')
DATE_FORMAT = getattr(cfg, 'DATE_FORMAT', '%Y-%m-%d')

CSV_DIALECT = getattr(cfg, 'CSV_DIALECT', 'excel')
CSV_DELIMITER = getattr(cfg, 'CSV_DELIMITER', None) or csv.get_dialect(CSV_DIALECT).delimiter

STAT_FILE = os.path.join(TEMP_DIR, f".{CFG_MODULE}.json")
GLUE_FILES = getattr(cfg, 'GLUE_FILES', True)

env = Environment(
    loader=FileSystemLoader([TEMPLATES_DIR, CFG_DIR]),
    autoescape=select_autoescape(['html', 'xml'])
)

JSON_TEMPLATE="""[
{%- for row in rows %}
{{"{"}}{% for k, v in zip(titles, row) %}"{{k}}": {% if v is not none %}{{v|tojson}}{% else %}null{% endif %}{% if not loop.last %}, {% endif %}{% endfor %}{{"}"}}{% if not loop.last %},{% endif %}
{%- endfor %}
]
"""
HTML_TEMPLATE="""<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta http-equiv="content-type" content="text/html; charset=UTF-8">
    <style type="text/css">
        th {{"{"}}background:lightblue; padding: 8px;{{"}"}}
        td {{"{"}}background:#e2e2e2; padding: 8px; text-align: left;{{"}"}}
    </style>
<title>{{title}}</title>
</head>
<body>
<h2>{{title}}</h2>
<table>
{%- if titles %}
<tr>{% for t in titles %}<th>{{t}}</th>{% endfor %}</tr>
{%- endif %}
{%- for row in rows %}
<tr>{% for d in row %}<td{{' style="text-align:right;"' if d is number else ''}}>{% if d is none %}{% elif d is number %}{{d|replace('.', dec_sep)}}{% else %}{{d}}{% endif %}</td>{% endfor %}</tr>
{%- endfor %}
</table>
</body>
</html>
"""

json_tpl = Template(JSON_TEMPLATE)
html_tpl = Template(HTML_TEMPLATE)


class RowsReader():

    def __init__(self, spec_name, spec, file_name, stat):
        self._spec_name = spec_name
        self._spec = spec
        self._file_name = file_name or self._spec.get('file', 'missing.file')
        self._stat = stat

        self._reader = None
        self._files = []
        self._files_index = -1
        self._file = None
        self._row_num = 0
        self._row_cnt = 0
        self._run_num = int(datetime.now().strftime('%Y%m%d%H%M%S'))
        self._recent_files = None
        self._latest_mtime = 0

    @staticmethod
    def naive_csv_reader(file, delimiter):
        for line in file:
            yield line.split(delimiter)

    def files(self):
        if self._files:
            return self._files

        self._file_path = self._file_name
        if not glob.glob(self._file_path):
            self._file_path = os.path.join(IN_DIR, self._file_path)
            if not glob.glob(self._file_path):
                logger.debug("%s - No files found: %s", self._spec_name, self._file_name)
                return self._files

        self._file_base = os.path.basename(self._file_path)
        self._file_ext = self._file_base.split('.')[-1]
        self.glue_files = self._spec.get('glue_files', GLUE_FILES)

        self._file = None
        self.wb = None

        # normalize the spec
        if self._spec.get('transformer') and not isinstance(self._spec['transformer'], (list, tuple)):
            self._spec['transformer'] = [self._spec['transformer']]

        # validate the spec
        assert self._file_ext in ('csv', 'xlsx', 'json', 'zip') or self._spec.get('text_lines', False), \
            f"Bad file extension \"{self._file_ext}\" in spec \"{self._spec_name}\""
        assert self._file_ext != 'json' or self._spec.get('transformer') or self._spec.get('text_lines', False), \
            f"Missing \"transformer\" in spec \"{self._spec_name}\""

        # find files to read
        all_files = []
        recent_files = []
        latest_mtime = self._stat[self._spec_name]['mtime']
        for file_name in sorted(glob.glob(self._file_path)):
            statinfo = os.stat(file_name)
            if statinfo.st_size == 0:
                logger.debug("%s - Skipping zero length file: %s", self._spec_name, file_name)
                continue
            all_files.append(file_name)
            file_mtime = statinfo.st_mtime
            if file_mtime > self._stat[self._spec_name]['mtime']:
                recent_files.append(file_name)
            latest_mtime = max(file_mtime, latest_mtime)

        if (not args.force and not self._spec.get('force') and not recent_files) or not all_files:
            logger.info("%s - No files found: %s", self._spec_name, self._file_name)
            return self._files

        if args.force or self._spec.get('force'):
            self._files = all_files
        else:
            self._files = recent_files
        self._recent_files = recent_files
        self._latest_mtime = latest_mtime

        logger.debug("%s - %s file(s) found: %s", self._spec_name, len(self._files), self._file_name)
        return self._files

    def _open_reader(self, file_path):
        _file = file_path
        if self._file_ext == 'zip':
            # unzip if zipfile contains the only file that meets the requirements
            import zipfile
            with zipfile.ZipFile(_file) as zf:
                inner_file = zf.namelist()
                assert len(inner_file) == 1, \
                    f"More than 1 member in file {_file}"
                inner_file = inner_file[0]
                assert os.path.basename(inner_file) == inner_file, \
                    f"Zip file member has path in file {_file}"
                inner_name, inner_ext = inner_file.rsplit('.', 1)
                ifile_path, ifile_name = os.path.split(_file)
                assert ifile_name.rsplit('.', 1)[0] in (inner_file, inner_name), \
                    f"Zip file member name is inconsistent with file name {_file}"
                assert inner_ext in ('csv', 'json', 'xlsx') or self._spec.get('text_lines', False), \
                    f"Zip file member has bad extension \"{inner_ext}\" in file {_file}"
                assert inner_ext != 'json' or self._spec.get('transformer'), \
                    f"Missing \"transformer\" in spec \"{self._spec_name}\""
                zf.extractall(ifile_path)
                extracted_file = os.path.join(ifile_path, inner_file)
            if os.path.isfile(extracted_file):
                if args.delete or self._spec.get('delete'):
                    os.remove(_file)
                _file = extracted_file
                _format = inner_ext
        else:
            _format = self._file_ext

        if self._spec.get('text_lines', False):
            self._file = open(_file, 'r', encoding=self._spec.get('encoding', ENCODING))
            self._reader = self._file
        elif _format == 'csv':
            self._file = open(_file, 'r', encoding=self._spec.get('encoding', ENCODING))
            if self._spec.get('csv_dialect', CSV_DIALECT) == 'naive':
                self._reader = \
                    self.naive_csv_reader(
                        self._file,
                        delimiter=self._spec.get('csv_delimiter', CSV_DELIMITER)
                    )
            else:
                self._reader = \
                    csv.reader(
                        self._file,
                        dialect=self._spec.get('csv_dialect', CSV_DIALECT),
                        delimiter=self._spec.get('csv_delimiter', CSV_DELIMITER)
                    )
        elif _format == 'xlsx':
            self._file = load_workbook(filename=_file, read_only=True)
            self._reader = self._file.worksheets[0].values
        elif _format == 'json':
            self._file = open(_file, 'r', encoding=self._spec.get('encoding', ENCODING))
            self._reader = (row for row in json.load(self._file))
            #assert isinstance(self._reader, list), f"Bad json file {_file}"
        logger.info(f"fi: {os.path.basename(_file)}")

    def _fetch_reader(self):
        if not self._reader:
            self._files_index += 1
            if len(self._files) > self._files_index:
                self._open_reader(self._files[self._files_index])
                self._row_num = 0
        logger.debug(f"_fetch_reader: {self._files[self._files_index] if self._reader else None}")

    def _close_reader(self):
        if self._reader:
            logger.debug(f"_close_reader: {self._files[self._files_index]}")
            self._reader = None
        if self._file:
            self._file.close()
            self._file = None

    def read(self, limit=0):
        logger.debug(f"read: limit={limit}")
        if limit == -1:
            self._close_reader()
        self._fetch_reader()
        rows = [] if self._reader else None
        while self._reader:
            for row in self._reader:
                self._row_cnt += 1
                self._row_num += 1
                if self._row_num <= self._spec.get('skip_lines', 0):
                    continue

                func = self._spec.get('transformer', [None])[0]
                row_data = \
                    func(row) if func and func.__code__.co_argcount == 1 else \
                    func(self._row_num, row) if func and func.__code__.co_argcount == 2 else \
                    func(self._run_num, self._row_cnt, row) if func and func.__code__.co_argcount == 3 else \
                    row if self._spec.get('text_lines', False) else \
                    row
                if isinstance(row_data, (list, tuple)) and row_data and not isinstance(row_data[0], (list, tuple)):
                    rows.append(row_data)
                elif isinstance(row_data, (list, tuple)) and row_data and isinstance(row_data[0], (list, tuple)):
                    rows.extend(row_data)
                elif row_data and self._spec.get('text_lines', False):
                    rows.append(row_data)
                #print(f"{self._row_cnt} : {self._row_num} : {row_data}")
                if self._row_num % 1000000 == 0:
                    logger.debug(f"{self._row_num // 1000000} mln rows")

                if limit > 0 and len(rows) == limit:
                    break # for

            if limit == -1:
                # all rows of a reader have been read, exit while loop
                self._close_reader()
            elif limit == 0:
                # all rows of a reader have been read
                self._close_reader()
                self._fetch_reader()
            elif rows in (None, []):
                # all rows of a reader have been read
                self._close_reader()
                self._fetch_reader()
            else:
                break # while
        logger.debug(f"read: {len(rows) if not rows is None else 'no'} rows")
        return rows

    def upstat(self):
        # set stat data no earlier than all files are successfully processed
        if self._recent_files:
            self._stat[self._spec_name]['mtime'] = self._latest_mtime


class RowsWriter():

    def __init__(self, spec_name, spec, file_name, reader):
        self._spec_name = spec_name
        self._spec = spec
        self._file_name = file_name or self._spec.get('file')
        self._reader = reader

    @staticmethod
    def _csv_row(row, dec_sep='.'):
        return tuple(
            '' if f == None else \
            f.strftime(DATETIME_FORMAT) if isinstance(f, datetime) else \
            f.strftime(DATE_FORMAT) if isinstance(f, date) else \
            str(f).replace('.', dec_sep) if isinstance(f, (float, complex, dec.Decimal)) else \
            str(f) \
            for f in row
            )

    @staticmethod
    def _jinja_row(row):
        return tuple(
            f.strftime(DATETIME_FORMAT) if isinstance(f, datetime) else \
            f.strftime(DATE_FORMAT) if isinstance(f, date) else \
            float(f) if isinstance(f, dec.Decimal) else \
            f
            for f in row
            )

    @staticmethod
    def _xlsx_row(row):
        # TypeError: Excel does not support timezones in datetimes. The tzinfo in the datetime/time object must be set to None.
        return tuple(
            f.replace(tzinfo=None) if isinstance(f, (time, datetime)) else \
            f
            for f in row
            )

    @staticmethod
    def _file_stem(stem, parts, seqn, user):
        filename_dict = dict()
        for part in parts:
            if part == 'date':
                filename_dict['date'] = date.today().isoformat()
            elif part == 'datetime':
                filename_dict['datetime'] = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            elif part == 'seqn':
                filename_dict['seqn'] = seqn
            elif part == 'user':
                filename_dict['user'] = user
        return stem % filename_dict

    @staticmethod
    def _finalize_file(filename, compress=False):
        _filename = filename[:-4]  # remove trailing '.out'
        if compress:
            import zipfile
            _zipfilename = _filename + '.zip'
            with zipfile.ZipFile(_zipfilename, 'w') as zf:
                zf.write(filename, arcname=os.path.basename(_filename))
            os.remove(filename)
            logger.info(f"fo: {os.path.basename(_zipfilename)}")
        else:
            # os.rename: On Windows, if dst exists a FileExistsError is always raised.
            if os.path.isfile(_filename):
                os.remove(_filename)
            os.rename(filename, _filename)
            logger.info(f"fo: {os.path.basename(_filename)}")

    def write(self):
        assert self._file_name, \
            f"Output file not specified, spec \"{self._spec_name}\""
        out_base, out_format = self._file_name.rsplit('.', 1)
        if out_format == 'zip':
            out_compress = True
            out_base, out_format = out_base.rsplit('.', 1)
        else:
            out_compress = False
        assert out_format in ('html', 'csv', 'xlsx', 'json') or \
            self._spec.get('text_lines', False) or (
            isinstance(self._spec.get('template'), str) and
            self._spec['template'].endswith(".jinja") and
            os.path.isfile(os.path.join(CFG_DIR, self._spec['template']))
            ), f"Bad format \"{out_format}\" or missing \"template\" in spec \"{self._spec_name}\""
        assert out_format != 'json' or \
            (self._spec.get('header') or self._spec.get('json.template')) or \
            self._spec.get('text_lines', False), \
            f"\"header\" not specified for json file with default template in spec \"{self._spec_name}\""

        out_path = self._spec.get('out_dir', OUT_DIR)
        rows_per_file = self._spec.get('rows_per_file', 0)
        encoding = self._spec.get(f"{out_format}.encoding", ENCODING)

        filename_parts = re.findall(r'%\((.+?)\)', out_base)
        assert not filename_parts or \
            set(filename_parts) <= {'date', 'datetime', 'seqn', 'user'}, \
            f"Bad named fields in filename: {self._spec['file']}"

        file = None
        seqn = 1  # start with 1
        rows = []
        if self._spec.get('text_lines', False):
            #
            # create output text file as is
            #
            while True:
                try:
                    rows = self._reader.read(rows_per_file)
                except:
                    if rows_per_file == -1 and self._spec['skip_bad_files']:
                        logger.exception('EXCEPT')
                        seqn += 1
                        continue
                    raise
                if rows is None:
                    # no file was read
                    break
                for row in rows:
                    # begin file
                    if file is None:
                        out_file = self._file_stem(out_base, filename_parts, seqn, args.user) + f".{out_format}.out"
                        seqn += 1
                        out_file = os.path.join(out_path, out_file)
                        file = open(out_file, 'w', encoding=encoding, errors='replace')
                        # no header
                    # next row
                    file.write(row)
                # end file
                if file:
                    file.close()
                    self._finalize_file(out_file, out_compress)
                    logger.info("fo: %s rows", len(rows))
                else:
                    logger.info("fo: nothing to write")
                file = None

        elif out_format in ('json', 'html'):
            #
            # create .json or .html file(s) using jinja2 template
            #
            template = self._spec.get(f"{out_format}.template", None)
            if template and os.path.isfile(os.path.join(TEMPLATES_DIR, template)):
                template = env.get_template(template)
            else:
                template = json_tpl if out_format == 'json' else html_tpl
            while True:
                try:
                    rows = self._reader.read(rows_per_file)
                except:
                    if rows_per_file == -1 and self._spec['skip_bad_files']:
                        logger.exception('EXCEPT')
                        seqn += 1
                        continue
                    raise
                if rows is None:
                    # no file was read
                    break
                if len(rows) > 0:
                    out_file = self._file_stem(out_base, filename_parts, seqn, args.user) + f".{out_format}.out"
                    seqn += 1
                    out_file = os.path.join(out_path, out_file)
                    file = open(out_file, 'w', encoding=encoding, errors='replace')
                    # header row
                    if not self._spec.get('header'):
                        self._spec['header'] = []
                    file.write(
                        template.render(
                            title=self._spec.get(f"{out_format}.title", self._spec_name),
                            titles=self._spec['header'] if self._spec.get('header') and self._spec.get(f"csv.header", True) or out_format == 'json' else [],
                            dec_sep=self._spec.get(f"{out_format}.dec_separator", '.'),
                            rows=[self._jinja_row(row) for row in rows],
                            zip=zip
                        )
                    )
                # end file
                if file:
                    file.close()
                    self._finalize_file(out_file, out_compress)
                    logger.info("fo: %s rows", len(rows))
                else:
                    logger.info("fo: nothing to write")
                file = None

        elif out_format == 'csv':
            #
            # create .csv file(s) line by line
            #
            dec_separator = self._spec.get('csv.dec_separator', '.')
            csv_dialect = self._spec.get('csv.dialect', CSV_DIALECT)
            csv_delimiter = self._spec.get('csv.delimiter', CSV_DELIMITER)

            while True:
                try:
                    rows = self._reader.read(rows_per_file)
                except:
                    if rows_per_file == -1 and self._spec['skip_bad_files']:
                        logger.exception('EXCEPT')
                        seqn += 1
                        continue
                    raise
                if rows is None:
                    # no file was read
                    break
                for row in rows:
                    # begin file
                    if file is None:
                        out_file = self._file_stem(out_base, filename_parts, seqn, args.user) + f".{out_format}.out"
                        seqn += 1
                        out_file = os.path.join(out_path, out_file)
                        file = open(out_file, 'w', encoding=encoding, errors='replace')
                        if csv_dialect == 'naive':
                            pass
                        else:
                            csv_writer = \
                                csv.writer(
                                    file,
                                    dialect=csv_dialect,
                                    delimiter=csv_delimiter,
                                    lineterminator='\n'
                                )
                        # header row
                        if self._spec.get('header') and self._spec.get('csv.header', True):
                            if csv_dialect == 'naive':
                                file.write(csv_delimiter.join(self._spec['header']) + '\n')
                            else:
                                csv_writer.writerow(self._spec['header'])
                    # next row
                    if csv_dialect == 'naive':
                        file.write(csv_delimiter.join(self._csv_row(row, dec_separator)) + '\n')
                    else:
                        csv_writer.writerow(self._csv_row(row, dec_separator))
                # end file
                if file:
                    file.close()
                    self._finalize_file(out_file, out_compress)
                    logger.info("fo: %s rows", len(rows))
                else:
                    logger.info("fo: nothing to write")
                file = None

        elif out_format == 'xlsx':
            #
            # create .xlsx file(s) row by row
            #
            while True:
                try:
                    rows = self._reader.read(rows_per_file)
                except:
                    if rows_per_file == -1 and self._spec['skip_bad_files']:
                        logger.exception('EXCEPT')
                        seqn += 1
                        continue
                    raise
                if rows is None:
                    # no file was read
                    break
                for row in rows:
                    # begin file
                    if file is None:
                        out_file = self._file_stem(out_base, filename_parts, seqn, args.user) + f".{out_format}.out"
                        seqn += 1
                        out_file = os.path.join(out_path, out_file)
                        file = Workbook(write_only=True)
                        sheet = file.create_sheet()
                        font = styles.Font(bold=True)
                        # header row
                        if self._spec.get('header') and self._spec.get('xlsx.header', True):
                            titles = [WriteOnlyCell(sheet, value=title) for title in self._spec['header']]
                            for title in titles:
                                title.font = font
                            sheet.append(titles)
                    # next row
                    sheet.append(self._xlsx_row(row))
                # end file
                if file:
                    file.save(out_file)
                    file.close()
                    self._finalize_file(out_file, out_compress)
                    logger.info("fo: %s rows", len(rows))
                else:
                    logger.info("fo: nothing to write")
                file = None


def process_spec(t):
    """
    Process spec from cfg-file.
    """
    spec_name, spec, in_file, out_file, stat = t
    return_code = 0
    logger.info(f"spec \"{spec_name}\"")
    try:
        spec['fo']['skip_bad_files'] = spec['fo'].get('skip_bad_files', spec['fi'].get('skip_bad_files', False))
        reader = RowsReader(spec_name, spec['fi'], in_file, stat)
        if reader.files():
            writer = RowsWriter(spec_name, spec['fo'], out_file, reader)
            writer.write()
        reader.upstat();
    except:
        logger.exception('EXCEPT')
        return_code = 1
    return return_code


def main():
    logger.info("-- start %s", ' '.join(sys.argv))
    _temp = datetime.now()
    logger.info(f"run with {PARALLEL_WORKERS} thread(s)")

    if os.path.isfile(STAT_FILE):
        with open(STAT_FILE, encoding='UTF-8') as f:
            stat = json.load(f)
    else:
        stat = dict()

    error_count = 0
    todo = []
    for spec_name, spec in specs.items():
        if spec_name not in stat:
            stat[spec_name] = {'mtime': 0}
        if SPEC in (spec_name, 'all', *spec.get('tags', [])):
            if spec.get('fi') is None or spec.get('fo') is None:
                logger.error('Skipping spec "%s" without "fi" and "fo" parts', spec_name)
                error_count += 1
                continue
            todo.append((spec_name, spec, args.fi, args.fo, stat))
            #error_count += process_spec(spec_name, spec, args.fi, args.fo, stat)

    with concurrent.futures.ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as ex:
        results = ex.map(process_spec, todo)
    error_count += sum(results)

    stat = {k: v for k, v in stat.items() if k in specs}
    with open(STAT_FILE, 'w', encoding='UTF-8') as f:
        f.write(json.dumps(stat))

    logger.debug(f"{datetime.now() - _temp}")
    logger.info("-- done%s", f" WITH {error_count} ERRORS" if error_count else '')
    sys.exit(error_count)


if __name__ == '__main__':
    main()
